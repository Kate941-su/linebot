import os
import openpyxl as px

event=os.environ["USER_ID"]
User_id=102
def make_or_open_file(EVENT):
    if os.path.isfile("./user"+str(EVENT)+".xlsx"):
        wb=px.load_workbook("./user"+str(EVENT)+".xlsx")#open xls file(wb=work book)
        print("exist")
    else:
        wb=px.Workbook()
        ws=wb.active
        ws.title="plan"
    wb.save("./user"+str(EVENT)+".xlsx")
    wb=px.load_workbook("./user"+str(EVENT)+".xlsx")#open xls file(wb=work book)
    ws = wb["plan"]#get sheet data(ws=work sheet)
    wb_w=px.load_workbook("./user"+str(EVENT)+".xlsx")
    ws_w=wb_w.worksheets[0]
    ws_w.cell(row=2,column=10,value=0)
    ws_w.cell(row=2,column=11,value=User_id)
    wb_w.save("./user"+str(EVENT)+".xlsx")
if __name__ ==  "__main__":
    make_or_open_file(event)
    
