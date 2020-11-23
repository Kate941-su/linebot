from flask import Flask, request, abort

import os,re

from linebot import LineBotApi, WebhookHandler

from linebot.exceptions import  InvalidSignatureError

from linebot.models import MessageEvent, TextMessage, TextSendMessage,TemplateSendMessage,ButtonsTemplate,MessageAction

from datetime import datetime
from datetime import timedelta

import openpyxl as px
from random import randint

import psycopg2 as p2
import psycopg2.extras
app = Flask(__name__)

#環境変数取得
LINE_CHANNEL_ACCESS_TOKEN = os.environ["LINE_CHANNEL_ACCESS_TOKEN"]
LINE_CHANNEL_SECRET = os.environ["LINE_CHANNEL_SECRET"]

line_bot_api = LineBotApi(LINE_CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(LINE_CHANNEL_SECRET)

@app.route("/")
def hello_world():
    return "hello what your name?"

@app.route("/kaito")
def aho():
    return str(10+20)


#Webhook設定
@app.route("/callback", methods=['POST'])
def callback():
    # get X-Line-Signature header value
    signature = request.headers['X-Line-Signature']

    # get request body as text
    body = request.get_data(as_text=True)
    app.logger.info("Request body: " + body)

    # handle webhook body
    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        abort(400)

    return 'OK'

#kokoまでは同じ

##メッセージ受信時
"""
@handler.add(MessageEvent, message=TextMessage)
def handle_message(event):
    text2=event.message.text
    line_bot_api.reply_message(
        event.reply_token,
#        TextSendMessage(text=event.message.text)
        TextSendMessage(text=text2),
        )


"""

@handler.add(MessageEvent, message=TextMessage)
def response_message(event):

#connect heroku postgresql
    Host=os.environ.get('PG_HOST')
    Port=os.environ.get('PG_PORT')
    Database=os.environ.get('PG_DBNM')
    User=os.environ.get('PG_USER')
    Password=os.environ.get('PG_PASS')
    connection = p2.connect("host="+str(Host)+" port="+str(Port)+" dbname="+str(Database)+" user="+str(User)+" password="+str(Password))
    connection.autocommit = True
    #これによってSQLオブジェクトを使用可能にする
    cur = connection.cursor()
    plan=1
    yyyy=2
    MM=3
    dd=4
    hh=5
    mm=6
    issue_id_col=9
    flag=10#列の名前のflag
    buffer1=11
    buffer2=12
    buffer3=13
    error_catch=14
    mistake=15
    send_id=16
    bffer_id=17
    buffer_date=18
    true_date=19
    true_time=20
    b_month=21
    b_day=22
    b_hour=23
    b_minute=24
#buffering_row
    b_row=201
    list_db=[]
    delete_db=[]
    list_id=[i for i in range(1,201)]
    list30=[4,6,9,11]
    list31=[1,3,5,7,8,10,12]

#入力ミス防止    
    error_flag=0
    Flag=0#条件分岐のためのflag
#    pattern = r'(0?[1-9]|1[0-2])[/\-月](0?[1-9]|[12][0-9]|3[01])日?$'#日付一致の正規表現
    profile = line_bot_api.get_profile(event.source.user_id)
    User_id=profile.user_id
    this_year=2020
    context = "{}"
    context = context.format("plan text,yyyy int,MM int,dd int,hh int,mmmm int,send_id text,issue_id int")
    cur.execute("create table if not exists User"+str(User_id)+"("+context+");")


#fileの有無　あればそれを開くなければつくってそれを開く
    if os.path.exists("./user"+str(User_id)+".xlsx"):
        wb=px.load_workbook("./user"+str(User_id)+".xlsx")#open xls file(wb=work book)
        ws = wb["plan"]#get sheet data(ws=work sheet)
        wb_w=px.load_workbook("./user"+str(User_id)+".xlsx")
        ws_w=wb_w.worksheets[0]
        ws_w.cell(row=b_row,column=send_id,value=User_id)
        wb_w.save("./user"+str(User_id)+".xlsx")
        wb=px.load_workbook("user"+str(User_id)+".xlsx")#open xls file(wb=work book)
        ws = wb["plan"]#get sheet data(ws=work sheet)
        Flag=int(ws.cell(row=b_row,column=flag).value)
        Mistake=int(ws.cell(row=b_row,column=mistake).value)
  
        if int(ws.cell(row=b_row,column=mistake).value)==3:
            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text="try again at first!!"),
            )
            ws_w.cell(row=b_row,column=mistake,value=0)
            ws_w.cell(row=b_row,column=flag,value=0)
    
    #Flag1 phase

        else:
            if Flag==0:
                if event.message.text == "予約":
                    line_bot_api.reply_message(
                    event.reply_token,
                    TextSendMessage(text="予約を行います日付を教えてください\nex)明日、今日、明後日、11/6\n※予約時刻は10分単位で行います1分単位で予約すると10分繰り上げ通知となります"),               
                )
                    ws_w.cell(row=b_row,column=flag,value=1)

                elif event.message.text == "ヘルプ":
                    line_bot_api.reply_message(
                    event.reply_token,
                    TextSendMessage(text="使い方を説明します！\n\n最初は、予約と入力してください！その後時刻と予定をこの会話に従って、入力してください！\n\n予定の確認をしたいときは、確認、削除したいときは削除と入力してください！\n\n予約を開始してからは確認、削除はできません。\n\n予約開始時以降３回入力ミスがあると自動的に最初からになるので注意してください。\n\nこの使い方をもう１回表示したいときは、ヘルプと入力してください！\n\n２００件を超えるリマインドは登録できないのでご注意ください。"),               
                )
                    ws_w.cell(row=b_row,column=flag,value=1)

                elif event.message.text == "確認":
                    cur.execute("select * from User"+str(User_id)+" order by issue_id DESC")#昇順
                #辞書型に格納したいがために新たなcurを定義

                    dictcur = connection.cursor(cursor_factory=p2.extras.DictCursor)
                    dictcur.execute("SELECT * FROM User"+str(User_id)+";")
                    result_dict=dictcur.fetchall()
                    dict_result = []
                #辞書型に格納
                    send_text="id　予定　月　日　時　分\n\n"
                    for row in result_dict:
                        dict_result.append(dict(row))
                    len_dic=len(result_dict)
                    for row in dict_result:
                        send_text+=str(row["issue_id"])+"　"+str(row["plan"])+"　"+str(row["mm"])+"月"+str(row["dd"])+"日　"+str(row["hh"])+"時"+str(row["mmmm"])+"分\n\n"
                    line_bot_api.reply_message(
                    event.reply_token,
                    TextSendMessage(text="予定の確認をします\n"+send_text),               
                )
                    ws_w.cell(row=b_row,column=flag,value=0)


                elif event.message.text == "削除":
                #辞書型に格納したいがために新たなcurを定義
                    dictcur = connection.cursor(cursor_factory=p2.extras.DictCursor)
                    dictcur.execute("SELECT * FROM User"+str(User_id)+";")
                    result_dict=dictcur.fetchall()
                    dict_result = []
                #辞書型に格納
                    send_text="id　予定　月　日　時　分\n\n"
                    for row in result_dict:
                        dict_result.append(dict(row))
                    len_dic=len(result_dict)
                    for row in dict_result:
                        send_text+=str(row["issue_id"])+"　"+str(row["plan"])+"　"+str(row["mm"])+"月"+str(row["dd"])+"日　"+str(row["hh"])+"時"+str(row["mmmm"])+"分\n\n"
                    line_bot_api.reply_message(
                    event.reply_token,
                    TextSendMessage(text="削除したいidをメッセージで送ってください。idは下の表から確認してください\n\n"+send_text),               
                )
                    ws_w.cell(row=b_row,column=flag,value=-1)

                else:
                    line_bot_api.reply_message(
                    event.reply_token,
                    TextSendMessage(text="不正な入力です！\nリマインダに予定を登録したい→予約\nリマインダの予定を確認したい→確認\nリマインダの予定を削除したい→削除"),
                )
                    ws_w.cell(row=b_row,column=flag,value=0)


#Flag-1削除フェーズ
            elif Flag==-1:
            #辞書型に格納したいがために新たなcurを定義
                dictcur = connection.cursor(cursor_factory=p2.extras.DictCursor)
                dictcur.execute("SELECT * FROM User"+str(User_id)+";")
                result_dict=dictcur.fetchall()
                dict_result = []
            #辞書型に格納
                for row in result_dict:
                    dict_result.append(dict(row))

                len_dic=len(result_dict)
                for i in range(len_dic):
                    list_db.append(dict_result[i]["issue_id"])
                for i in range(len_dic):
                    delete_db.append(dict_result[i]["issue_id"])
        
                delete_id=event.message.text
                try :
                    ids=int(delete_id)                
                    if ids in list_db:
                        line_bot_api.reply_message(
                        event.reply_token,
                        TextSendMessage(text="削除しました"),               
                    )                
                        cur.execute("delete from User"+str(User_id)+" where issue_id="+str(ids))   
                        ws_w.cell(row=b_row,column=flag,value=0)     
                    else:
                        line_bot_api.reply_message(
                        event.reply_token,
                        TextSendMessage(text="そのようなデータはありません。削除したいidをメッセージで送ってください。\n\nhelloworld")          
                    ) 
                        ws_w.cell(row=b_row,column=flag,value=-1)
                        ws_w.cell(row=b_row,column=mistake,value=Mistake+1)   
                except:
                        line_bot_api.reply_message(
                        event.reply_token,
                        TextSendMessage(text="そのようなデータはありません。削除したいidをメッセージで送ってください。\n\n"+str(ids)),                     
                    )                       
                        ws_w.cell(row=b_row,column=flag,value=-1)
                        ws_w.cell(row=b_row,column=mistake,value=Mistake+1)

            elif Flag==1:
                is_message_date = event.message.text
                ws_w.cell(row=b_row,column=buffer1,value=is_message_date)
                wb_w.save("user"+str(User_id)+".xlsx")
                wb=px.load_workbook("user"+str(User_id)+".xlsx")#open xls file(wb=work book)
                ws = wb["plan"]#get sheet data(ws=work sheet)
                #型を判定する
                #datetime型で方が一致していた時
                if str(ws.cell(row=b_row,column=buffer1).value) == "今日":   
                    line_bot_api.reply_message(
                    event.reply_token,
                    TextSendMessage(text="何時何分に設定しますか\n入力フォーマット例(11時00分のとき):11:00（半角）\n※10分単位です!"),
                    )
                    ws_w.cell(row=b_row,column=flag,value=2)
                    ws_w.cell(row=b_row,column=buffer1,value=event.message.text)#issue id

                elif str(ws.cell(row=b_row,column=buffer1).value) == "明日":   
                    line_bot_api.reply_message(
                    event.reply_token,
                    TextSendMessage(text="何時何分に設定しますか\n入力フォーマット例(11時10分のとき):11:10（半角）"),
                    )
                    ws_w.cell(row=b_row,column=flag,value=2)
                    ws_w.cell(row=b_row,column=buffer1,value=event.message.text)#issue id

                elif str(ws.cell(row=b_row,column=buffer1).value) == "明後日":   
                    line_bot_api.reply_message(
                    event.reply_token,
                    TextSendMessage(text="何時何分に設定しますか\n入力フォーマット例(11時10分のとき):11:10（半角）"),
                    )
                    ws_w.cell(row=b_row,column=flag,value=2)
                    ws_w.cell(row=b_row,column=buffer1,value=event.message.text)#issue id

                    #今日明日明後日意外の処理
                else:                
                    r_message=event.message.text           
                    try:
                        k=0
                        r_message=r_message.split("/")
                        for i in r_message:
                            r_message[k]=int(i)
                            k+=1
                        if r_message[0] in list30:
                            if r_message[1]>30:
                                error_flag=1
                        elif r_message[0] in list31:
                            if r_message[1]>31:
                                error_flag=1
                        else:
                            if r_message[1]>28:
                                error_flag=1
                    except:
                        error_flag=1


                    if error_flag == 0:    
                        line_bot_api.reply_message(
                        event.reply_token,
                        TextSendMessage(text="何時何分に設定しますか\n入力フォーマット例(11時10分のとき):11:10（半角）"),
                        )
                        ws_w.cell(row=b_row,column=flag,value=2)
                        ws_w.cell(row=b_row,column=b_month,value=r_message[0])#issue id
                        ws_w.cell(row=b_row,column=b_day,value=r_message[1])
                    else:
                        line_bot_api.reply_message(
                        event.reply_token,
                        TextSendMessage(text="入力ミスがあります。ex)明日、今日、明後日、11/6"),
                        )
                        ws_w.cell(row=b_row,column=flag,value=1)
        #               ws_w.cell(row=b_row,column=buffer1,value=event.message.text)
                        ws_w.cell(row=b_row,column=mistake,value=Mistake+1)

    #Flag2 phase

            elif Flag==2:
                r_message=event.message.text           
                try:
                    k=0
                    r_message=r_message.split(":")
                    for i in r_message:
                        r_message[k]=int(i)
                        k+=1
                    if r_message[0]>=25 or r_message[1]>=60:
                        error_flag=1
                except:
                    error_flag=1


                if error_flag == 0:    
                    line_bot_api.reply_message(
                    event.reply_token,
                    TextSendMessage(text="何の予定がありますか？\n"),
                    )
                    ws_w.cell(row=b_row,column=flag,value=3)
                    ws_w.cell(row=b_row,column=b_hour,value=r_message[0])#issue id
                    ws_w.cell(row=b_row,column=b_minute,value=r_message[1])#issue id
                else:
                    line_bot_api.reply_message(
                    event.reply_token,
                    TextSendMessage(text="入力ミスがあります。\n何時何分に設定しますか\n入力フォーマット例(11時10分のとき):11:10（半角）"),
                    )
                    ws_w.cell(row=b_row,column=flag,value=2)
                    ws_w.cell(row=b_row,column=mistake,value=Mistake+1)
    #Flag3 phase

            elif Flag == 3:
                ws_w.cell(row=b_row,column=buffer3,value=event.message.text)#issue id
                ws_w.cell(row=b_row,column=flag,value=0)
                ws_w.cell(row=b_row,column=send_id,value=profile.user_id)
                wb_w.save("user"+str(User_id)+".xlsx")
                wb=px.load_workbook("user"+str(User_id)+".xlsx")#open xls file(wb=work book)
                ws = wb["plan"]#get sheet data(ws=work sheet)

            #辞書型に格納したいがために新たなcurを定義
                dictcur = connection.cursor(cursor_factory=p2.extras.DictCursor)
                dictcur.execute("SELECT * FROM User"+str(User_id)+";")
                result_dict=dictcur.fetchall()
                dict_result = []
            #辞書型に格納
                for row in result_dict:
                    dict_result.append(dict(row))
                len_dic=len(result_dict)
                for i in range(len_dic):
                    list_db.append(dict_result[i]["issue_id"])
#   もしデータベースに予約がなかったらissue_id=1あったらissue_id
                if len_dic == 0:
                    issue_id=1
                else:
#                    issue_id=int(dict_result[len_dic-1]["issue_id"])+1
                    for i in list_id:
                        if i in list_db:
                            pass
                        else:
                            issue_id=i
                            break
                #予定の処理
                ws_w.cell(row=issue_id,column=plan,value=ws.cell(row=b_row,column=buffer3).value)
                #普通に日付を入れたときの処理
                ws_w.cell(row=issue_id,column=yyyy,value=this_year)
                ws_w.cell(row=issue_id,column=MM,value=ws.cell(row=b_row,column=b_month).value)
                ws_w.cell(row=issue_id,column=dd,value=ws.cell(row=b_row,column=b_day).value)
                ws_w.cell(row=issue_id,column=hh,value=ws.cell(row=b_row,column=b_hour).value)
                ws_w.cell(row=issue_id,column=mm,value=ws.cell(row=b_row,column=b_minute).value)
                dt=datetime.now()
                dt+=timedelta(hours=9)
                month=dt.month
                day=dt.day
                #今日明日明後日の処理
               #日本時間調整

#                english_hour=datetime.now().hour
#                japanese_hour=english_hour+9
#                if japanese_hour >=24:
#                    day+=1
#                    if month in list30:
#                        if day>30:
#                            day-=30
#                            month+=1
#                    if month in list31:
#                        if day > 31:
#                            day-=31
#                            month+=1
#                    else:
#                        if day>27:
#                            day-=27
#                            month+=1
                
                if ws.cell(row=b_row,column=buffer1).value=="今日":
                    ws_w.cell(row=issue_id,column=MM,value=month)
                    ws_w.cell(row=issue_id,column=dd,value=day)

                #tomorrow
                elif ws.cell(row=b_row,column=buffer1).value == "明日":
                    ws_w.cell(row=issue_id,column=MM,value=month)#month
                    tomorrow=day+1
                    ws_w.cell(row=issue_id,column=dd,value=tomorrow)
                    #save and reopen
                    wb_w.save("user"+str(User_id)+".xlsx")
                    wb=px.load_workbook("user"+str(User_id)+".xlsx")#reopen xls file(wb=work book)
                    ws = wb["plan"]#get sheet data(ws=work sheet)

                    #月繰り上げ処理:list30は30日まで、list31は31日まで
                    if int(ws.cell(row=issue_id,column=MM).value) in list30:
                        if tomorrow>30:
                            ws_w.cell(row=issue_id,column=dd,value=tomorrow-30)
                            Month=ws.cell(row=issue_id,column=MM).value
                            ws_w.cell(row=issue_id,column=MM,value=Month+1)
                    
                    elif int(ws.cell(row=issue_id,column=MM).value) in list31:
                        if tomorrow>31:
                            ws_w.cell(row=issue_id,column=dd,value=tomorrow-31)
                            Month=ws.cell(row=issue_id,column=MM).value
                            ws_w.cell(row=issue_id,column=MM,value=Month+1)
                    #2月の処理
                    else:
                        if tomorrow>28:
                            ws_w.cell(row=issue_id,column=dd,value=day-28)
                            Month=ws.cell(row=issue_id,column=MM).value
                            ws_w.cell(row=issue_id,column=MM,value=Month+1)



                #day after tomorrow
                elif ws.cell(row=b_row,column=buffer1).value == "明後日":
                    ws_w.cell(row=issue_id,column=MM,value=month)#month
                    d_a_tomorrow=day+2
                    ws_w.cell(row=issue_id,column=dd,value=d_a_tomorrow)
                    #save and reopen
                    wb_w.save("user"+str(User_id)+".xlsx")
                    wb=px.load_workbook("user"+str(User_id)+".xlsx")#reopen xls file(wb=work book)
                    ws = wb["plan"]#get sheet data(ws=work sheet)

                    #月繰り上げ処理
                    if int(ws.cell(row=issue_id,column=MM).value) in list30:
                        if d_a_tomorrow>29:
                            ws_w.cell(row=issue_id,column=dd,value=d_a_tomorrow-30)
                            Month=ws.cell(row=issue_id,column=MM).value
                            ws_w.cell(row=issue_id,column=MM,value=Month+1)
                    
                    elif int(ws.cell(row=issue_id,column=MM).value) in list31:
                        if d_a_tomorrow>30:
                            ws_w.cell(row=issue_id,column=dd,value=d_a_tomorrow-31)
                            Month=ws.cell(row=issue_id,column=MM).value
                            ws_w.cell(row=issue_id,column=MM,value=Month+1)
                    #2月の処理
                    else:
                        if d_a_tomorrow>27:
                            ws_w.cell(row=issue_id,column=dd,value=d_a_tomorrow-28)
                            Month=ws.cell(row=issue_id,column=MM).value
                            ws_w.cell(row=issue_id,column=MM,value=Month+1)


                wb_w.save("user"+str(User_id)+".xlsx")
                wb=px.load_workbook("user"+str(User_id)+".xlsx")#reopen xls file(wb=work book)
                ws = wb["plan"]#get sheet data(ws=work sheet)
                
                #年明け処理
                if int(ws.cell(row=issue_id,column=MM).value) >12:
                    ws_w.cell(row=issue_id,column=MM,value=1)
                    ws_w.cell(row=issue_id,column=yyyy,value=this_year+1)

                wb_w.save("user"+str(User_id)+".xlsx")
                wb=px.load_workbook("user"+str(User_id)+".xlsx")#open xls file(wb=work book)
                ws = wb["plan"]#get sheet data(ws=work sheet)

                if datetime.now() > datetime(year=this_year,month=ws.cell(row=issue_id,column=MM).value,day=ws.cell(row=issue_id,column=dd).value,hour=ws.cell(row=issue_id,column=hh).value,minute=ws.cell(row=issue_id,column=mm).value):
                    ws_w.cell(row=issue_id,column=yyyy,value=this_year+1)
                else:
                    ws_w.cell(row=issue_id,column=yyyy,value=this_year)



                wb_w.save("user"+str(User_id)+".xlsx")
                wb=px.load_workbook("user"+str(User_id)+".xlsx")#open xls file(wb=work book)
                ws = wb["plan"]#get sheet data(ws=work sheet)

                Plan=ws.cell(row=issue_id,column=plan).value
                Year=int(ws.cell(row=issue_id,column=yyyy).value)
                Month=int(ws.cell(row=issue_id,column=MM).value)
                Day=int(ws.cell(row=issue_id,column=dd).value)
                Hour=int(ws.cell(row=issue_id,column=hh).value)
                Minute=int(ws.cell(row=issue_id,column=mm).value)


                Send_id=str(User_id)
                Issue_id=issue_id

                cur.execute("insert into User"+str(User_id)+" values(%s,%s,%s,%s,%s,%s,%s,%s);",(Plan,Year,Month,Day,Hour,Minute,Send_id,Issue_id))
                cur.execute("select * from User"+str(User_id)+" order by issue_id")
                if issue_id == 200:
                    line_bot_api.reply_message(
                    event.reply_token,
                    TextSendMessage(text="予約上限です\n削除するか通知を待ってから追加できます"),        
                    )
                    
                else:
                    if type(ws.cell(row=b_row,column=buffer1).value) == type("hello"):#buffer1の内容はb_rowからissue_idに移していない
                        line_bot_api.reply_message(
                        event.reply_token,
                        TextSendMessage(text=str(ws.cell(row=b_row,column=buffer1).value)+"の"+str(ws.cell(row=issue_id,column=hh).value)+"時"+str(ws.cell(row=issue_id,column=mm).value)+"分に"+"”"+str(ws.cell(row=b_row,column=buffer3).value)+"”"+"で予約しました。\n"),        
                        )
                    else:
                        line_bot_api.reply_message(
                        event.reply_token,
                        TextSendMessage(text=str(ws.cell(row=issue_id,column=yyyy).value)+"年"+str(ws.cell(row=issue_id,column=MM).value)+"月"+str(ws.cell(row=issue_id,column=dd).value)+"日の"+str(ws.cell(row=issue_id,column=hh).value)+"時"+str(ws.cell(row=issue_id,column=mm).value)+"分に"+"”"+str(ws.cell(row=b_row,column=buffer3).value)+"”"+"で予約しました。\n"),        
                        )



#                    if ws.cell(row=b_row,column=b_month).value != 0:
#                        line_bot_api.reply_message(
#                        event.reply_token,
#                        TextSendMessage(text=str(ws.cell(row=issue_id,column=yyyy).value)+"年"+str(ws.cell(row=issue_id,column=MM).value)+"月"+str(ws.cell(row=issue_id,column=dd).value)+"日の"+str(ws.cell(row=issue_id,column=hh).value)+"時"+str(ws.cell(row=issue_id,column=mm).value)+"分に"+"”"+str(ws.cell(row=b_row,column=buffer3).value)+"”"+"で予約しました。\n"+str(issue_id)),        
#                        )
#                    else:
#                        line_bot_api.reply_message(
#                        event.reply_token,
#                        TextSendMessage(text=str(ws.cell(row=b_row,column=buffer1).value)+"の"+str(ws.cell(row=b_row,column=b_hour).value)+"時"+str(ws.cell(row=b_row,column=b_minute).value)+"分に"+"”"+str(ws.cell(row=b_row,column=buffer3).value)+"”"+"で予約しました。\n"+str(issue_id)),        
#                        )





    #            ws_w.cell(row=b_row,column=buffer3,value=event.message.text)#issue id
    #            ws_w.cell(row=b_row,column=flag,value=0) 
        
        wb_w.save("user"+str(User_id)+".xlsx")
        connection.close()

    else:
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text="登録までお待ちください"),
        ) 
        user_id = os.environ["MY_ID"]
        line_bot_api.push_message(user_id, TextSendMessage(text=str(datetime.now().month)+"月"+str(datetime.now().day)+"日\n"+"user"+str(User_id)+"さんが登録を要請しました。\nファイルの作成をしてください。\nファイル名 user"+str(User_id)+".xlsx")
        )



#error時の対応

#sample bot
"""
    profile = line_bot_api.get_profile(event.source.user_id)

    status_msg = profile.status_message
    if status_msg != "None":
        # LINEに登録されているstatus_messageが空の場合は、"なし"という文字列を代わりの値とする
        status_msg = "なし"

    messages = TemplateSendMessage(alt_text="Buttons template",
                                   template=ButtonsTemplate(
                                       thumbnail_image_url=profile.picture_url,
                                       title=profile.display_name,
                                       text=f"User Id: {profile.user_id[:8]}...\n"
                                            f"Status Message: {status_msg}",
                                       actions=[MessageAction(label="成功", text="次は何を実装しましょうか？")]))

    line_bot_api.reply_message(event.reply_token, messages=messages)
"""


if __name__ == "__main__":
#    app.run()
    port = int(os.getenv("PORT"))
    app.run(host="0.0.0.0", port=port)