import openpyxl as px
import pprint
from random import randint
import re
from datetime import date,datetime,timedelta

from linebot import LineBotApi
from linebot.models import TextSendMessage
from linebot.exceptions import LineBotApiError
import os

#today's infomation
year=datetime.now().year
month=datetime.now().month
today=datetime.now().day
hour=datetime.now().hour
minute=datetime.now().minute


your_file_list=["./sample.xlsx","./sample2.xlsx"]

list30=[4,6,9,11]
list31=[1,3,5,7,8,10,12]
#Max_issue_id
max_issue_id=100

plan=1
yyyy=2
MM=3
dd=4
hh=5
mm=6
issue_id_col=9
flag=10#列の名前のflag
buffer1=11
buffer2=12
buffer3=13
error_catch=14
mistake=15
send_id=16
buffer_id=17
buffer_date=18
true_date=19
true_time=20


for your_file in your_file_list: 

    wb=px.load_workbook(your_file)#open xls file(wb=work book)
    ws = wb["plan"]#get sheet data(ws=work sheet)
    col_end=ws.max_column
    row_end=ws.max_row
    #print(row_end,col_end)
    id=ws.cell(row_end-1,col_end).value
    #id=int(id)#cast float -> int
    buffer_col=7

    #write in text
    wb_w=px.load_workbook(your_file)
    ws_w=wb_w.worksheets[0]
    #reread in text

    #本来ならrow=issue_id
    ws_w.cell(row=2,column=yyyy).value=year
    ws_w.cell(row=5,column=buffer1,value="11月3日")
    #試し
    #ws_w.cell(row=5,column=buffer2,value="11:11")
    wb.save(your_file)
    wb=px.load_workbook(your_file)#open xls file(wb=work book)
    ws = wb["plan"]#get sheet data(ws=work sheet)
    print(type(ws.cell(row=10,column=buffer1).value))
    #flag1 phase
    #today
    #確認
    #print(type(ws.cell(row=5,column=buffer1).value))
    #print(ws.cell(row=2,column=buffer1).value)
    print(type(ws.cell(row=5,column=buffer1).value) is datetime)




    if ws.cell(row=2,column=buffer1).value is datetime:
        try:
            obj=ws.cell(row=2,column=buffer1).value
            obj_month=obj.month
            obj_day=obj.day
            ws_w.cell(row=2,column=MM,value=obj_month)
            ws_w.cell(row=2,column=dd,value=obj_day)
            wb_w.save(your_file)
        except:
            print("error occured!!\nphase1")
            ws_w.cell(row=2,column=error_catch,value=1)
            wb_w.save(your_file)



    #flag2 phase
    try :
        obj=ws.cell(row=2,column=buffer2).value
        obj_hour=obj.hour
        obj_min=obj.minute
        ws_w.cell(row=2,column=hh,value=obj_hour)
        ws_w.cell(row=2,column=mm,value=obj_min)
        wb_w.save(your_file)
        #save and reopen
        wb=px.load_workbook(your_file)#reopen xls file(wb=work book)
        ws = wb["plan"]#get sheet data(ws=work sheet)
        #上の二行でリロードしてからtestを行う
        test=int(ws.cell(row=2,column=mm).value)

    #エラー発生時の振り出しに戻す対応
    except:
        print("error occured!!\nphase2")
        ws_w.cell(row=2,column=error_catch,value=1)
        wb_w.save(your_file)   

    #flag3 phase
    Plan=ws_w.cell(row=2,column=buffer3).value
    ws_w.cell(row=2,column=plan,value=Plan)
    ws_w.cell(row=8,column=plan,value='=SUM(A1:A3)')
    wb_w.save(your_file) 

    if datetime(ws.cell(row=2,column=yyyy).value,ws.cell(row=2,column=MM).value,ws.cell(row=4,column=dd).value,ws.cell(row=4,column=hh).value,ws.cell(row=4,column=mm).value) < datetime.now():

        LINE_CHANNEL_ACCESS_TOKEN = os.environ["LINE_CHANNEL_ACCESS_TOKEN"]

        line_bot_api = LineBotApi(LINE_CHANNEL_ACCESS_TOKEN)

        #print(ws.cell(row=2,column=buffer1).value.month)
        user_id = ws.cell(row=3,column=send_id).value
#        if (type(ws.cell(row=3,column=buffer1).value) is datetime):
#            message = TextSendMessage(text=str(ws.cell(row=3,column=buffer1).value.month)+"月"+str(ws.cell(row=3,column=buffer1).value.day)+"日"+"に"+str(ws.cell(row=3,column=buffer3).value)+"で予約しました。\n"+str(ws.cell(row=3,column=send_id).value))
#            line_bot_api.push_message(user_id,message)
#        else:
#            message = TextSendMessage(text=str(ws.cell(row=3,column=buffer1).value)+"の"+str(ws.cell(row=3,column=buffer2).value)+"に"+str(ws.cell(row=3,column=buffer3).value)+"で予約しました。\n"+str(ws.cell(row=3,column=send_id).value))
#            line_bot_api.push_message(user_id,message)

        for i in range(2,max_issue_id+1):
            is_issueid = ws.cell(row=i,column=issue_id_col).value
            if is_issueid != None:
                if datetime(ws.cell(row=i,column=yyyy).value,ws.cell(row=i,column=MM).value,ws.cell(row=i,column=dd).value,ws.cell(row=i,column=hh).value,ws.cell(row=i,column=mm).value) < datetime.now():
                    message = TextSendMessage(text=str(ws.cell(row=i,column=buffer3).value)+"の時間です！")
                    line_bot_api.push_message(user_id,message)
